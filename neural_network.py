import openpyxl

workbook_bitcoin = openpyxl.load_workbook('bitcoin_data.xlsx')
workbook_video_games = openpyxl.load_workbook('shippment_data.xlsx')
workbook_shippment = openpyxl.load_workbook('video_game_sales.xlsx')

print(workbook_bitcoin)